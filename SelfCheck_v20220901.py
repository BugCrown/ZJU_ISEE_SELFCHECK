#-------- encoding: utf-8 --------
'''
@File    :   SelfCheck_v20220901.py
@Time    :   2022/09/13 17:03:41
@Author  :   BugCrown 
@Version :   2209a
@Contact :   grantbugcrown@gmail.com
'''
#---------------------------------

from openpyxl  import load_workbook
import os
import re

# The sheet you will install from the colledge's website
print("想要打开的文件(xls/xlsx):")
xlsx = input()
wb1 = load_workbook('./' + xlsx)

ws1 = wb1['Sheet1']

course_name = []
course_id = []
course_weight = []
course_grade = []
course_status = []

for i in range(7, 250):
    course_name.append(str(ws1.cell(row = i, column = 2).value).strip())
    course_id.append(str(ws1.cell(row = i, column = 4).value).strip())
    course_weight.append(str(ws1.cell(row = i, column = 8).value).strip())
    course_grade.append(str(ws1.cell(row = i, column = 9).value).strip())
    course_status.append(str(ws1.cell(row = i, column = 10).value).strip())

# The sheet you should fill
wb2 = load_workbook('./祝桢涛.xlsx')

ws2 = wb2['Sheet1']

for i in range(7, 250):
    for j in range(7, 250):
        if str(ws2.cell(row = i, column = 3).value).strip() == course_name[j - 7]:
            if course_status[j - 7] == '通过':
                d = ws2.cell(row = i, column = 5, value = course_grade[j - 7])
                break
            else:
                break
            
# Personal course and interprofessional course
pcourse_name = []
pcourse_id = []
pcourse_weight = []
pcourse_grade = []
pcourse_status = []
icourse_name = []
icourse_id = []
icourse_weight = []
icourse_grade = []
icourse_status = []

for i in range(7, 250):
    ret = re.search(r"个性", course_name[i - 7])
    if ret != None:
        while i < 250:
            i = i + 1
            ret = re.search(r"跨专业", course_name[i - 7])
            if ret == None:
                pcourse_name.append(course_name[i - 7])
                pcourse_id.append(course_id[i - 7])
                pcourse_weight.append(course_weight[i - 7])
                pcourse_grade.append(course_grade[i - 7])
                pcourse_status.append(course_status[i - 7])
            else:
                break

for i in range(7, 250):
    ret = re.search(r"个性", str(ws2.cell(row = i, column = 1).value))
    if ret != None:
        j = 0
        k = 1
        while j < len(pcourse_name):
            if pcourse_status[j] == '通过':
                nm = ws2.cell(row = i + k, column = 3, value = pcourse_name[j])
                id = ws2.cell(row = i + k, column = 2, value = pcourse_id[j])
                wg = ws2.cell(row = i + k, column = 4, value = pcourse_weight[j])
                gd = ws2.cell(row = i + k, column = 5, value = pcourse_grade[j])
                if gd.value == 'None':
                    gd = ws2.cell(row = i + k, column = 5, value = '通过')
                k = k + 1
                j = j + 1
            else:
                j = j + 1

for i in range(7, 250):
    ret = re.search(r"跨专业", course_name[i - 7])
    if ret != None:
        while i < 250:
            i = i + 1
            ret = re.search(r"国际", course_name[i - 7])
            if ret == None:
                icourse_name.append(course_name[i - 7])
                icourse_id.append(course_id[i - 7])
                icourse_weight.append(course_weight[i - 7])
                icourse_grade.append(course_grade[i - 7])
                icourse_status.append(course_status[i - 7])
            else:
                break

for i in range(7, 250):
    ret = re.search(r"跨专业", str(ws2.cell(row = i, column = 1).value))
    if ret != None:
        j = 0
        k = 1
        while j < len(icourse_name):
            if pcourse_status[j] == '通过':
                nm = ws2.cell(row = i + k, column = 3, value = icourse_name[j])
                id = ws2.cell(row = i + k, column = 2, value = icourse_id[j])
                wg = ws2.cell(row = i + k, column = 4, value = icourse_weight[j])
                gd = ws2.cell(row = i + k, column = 5, value = icourse_grade[j])
                if gd.value == 'None':
                    gd = ws2.cell(row = i + k, column = 5, value = '通过')
                k = k + 1
                j = j + 1
            else:
                j = j + 1

print("另存为(带上后缀名xls/xlsx):")
save = input()
wb2.save(save)