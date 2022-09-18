#-------- encoding: utf-8 --------
'''
@File    :   SelfCheck.py
@Time    :   2022/09/18 11:01:45
@Author  :   BugCrown 
@Version :   Formal
@Contact :   grantbugcrown@gmail.com
'''
#---------------------------------

from tabnanny import check
from openpyxl  import load_workbook
import os
import re

# A class includes every class's information
class course:
    def __init__(self, name, id, weight, grade, status):
        self.name = name
        self.id = id
        self.weight = weight
        self.grade = grade
        self.status = status

def begin_end(bgn: str, end: str, name: list):
    bgn_num = 0
    end_num = 0
    for i in range(len(name)):
        ret0 = re.search(bgn, name[i])
        ret1 = re.search(end, name[i])
        if ret0 != None:
            bgn_num = i
        if ret1 != None:
            end_num = i
    return [bgn_num, end_num]

def check_and_fill(bgn: str, end: str, class0: course, class1: course):
    [bgn_num, end_num] = begin_end(bgn, end, class0.name)
    course_num = 0
    for i in range(bgn_num, end_num):
        if class0.status[i] == '通过':
            course_num = course_num + 1
            class1.name.append(class0.name[i])
            class1.id.append(class0.id[i])
            class1.weight.append(class0.weight[i])
            class1.grade.append(class0.grade[i])
            class1.status.append(class0.status[i])
    return course_num
# The sheet you will install from the colledge's website
print("教务系统文件(xls/xlsx):")
xlsx1 = input()
wb1 = load_workbook('./' + xlsx1)

ws1 = wb1['Sheet1']

all = course([], [], [], [], [])

for i in range(7, 250):
    all.name.append(str(ws1.cell(row = i, column = 2).value).strip())
    all.id.append(str(ws1.cell(row = i, column = 4).value).strip())
    all.weight.append(str(ws1.cell(row = i, column = 8).value).strip())
    all.grade.append(str(ws1.cell(row = i, column = 9).value).strip())
    all.status.append(str(ws1.cell(row = i, column = 10).value).strip())

# The sheet you should fill
print("信电文件:")
xlsx2 = input()
wb2 = load_workbook('./' + xlsx2)

ws2 = wb2['Sheet1']

for i in range(7, 250):
    for j in range(7, 250):
        if str(ws2.cell(row = i, column = 3).value).strip() == all.name[j - 7]:
            if all.status[j - 7] == '通过':
                d = ws2.cell(row = i, column = 5, value = all.grade[j - 7])
                break
            else:
                break
            
# Personalized course and interprofessional course
per = course([], [], [], [], [])
inter = course([], [], [], [], [])
per_num = check_and_fill(r"个性", r"跨专业", all, per)
inter_num = check_and_fill(r"跨专业", r"国际化", all, inter)

for i in range(7, 250):
    ret = re.search(r"个性", str(ws2.cell(row = i, column = 1).value))
    if ret != None:
        j = 0
        k = 1
        while j < len(per.name):
            if per.status[j] == '通过':
                nm = ws2.cell(row = i + k, column = 3, value = per.name[j])
                id = ws2.cell(row = i + k, column = 2, value = per.id[j])
                wg = ws2.cell(row = i + k, column = 4, value = per.weight[j])
                gd = ws2.cell(row = i + k, column = 5, value = per.grade[j])
                if gd.value == 'None':
                    gd = ws2.cell(row = i + k, column = 5, value = '通过')
                k = k + 1
                j = j + 1
            else:
                j = j + 1

for i in range(7, 250):
    ret = re.search(r"跨专业", str(ws2.cell(row = i, column = 1).value))
    if ret != None:
        j = 0
        k = 1
        while j < len(inter.name):
            if inter.status[j] == '通过':
                nm = ws2.cell(row = i + k, column = 3, value = inter.name[j])
                id = ws2.cell(row = i + k, column = 2, value = inter.id[j])
                wg = ws2.cell(row = i + k, column = 4, value = inter.weight[j])
                gd = ws2.cell(row = i + k, column = 5, value = inter.grade[j])
                if gd.value == 'None':
                    gd = ws2.cell(row = i + k, column = 5, value = '通过')
                k = k + 1
                j = j + 1
            else:
                j = j + 1

# Second/Third/Fourth Courses
other_courses = [0, 0, 0, 0]

for i in range(7, 250):
    ret = re.search(r"不通过", all.name[i - 7])
    if re.search(r"第二课堂<必修课程>", all.name[i - 7]) != None:
        if ret == None:
            other_courses[0] = 1
    if re.search(r"第三课堂<必修课程>", all.name[i - 7]) != None:
        if ret == None:
            other_courses[1] = 1
    if re.search(r"第四课堂<必修课程>", all.name[i - 7]) != None:
        if ret == None:
            other_courses[2] = 1
    if re.search(r"国际化模块<选修课程>", all.name[i - 7]) != None:
        if ret == None:
            other_courses[3] = 1

for i in range(7, 250):
    ret0 = re.search(r"第二课堂", str(ws2.cell(row = i, column = 1).value))
    ret1 = re.search(r"第三课堂", str(ws2.cell(row = i, column = 1).value))
    ret2 = re.search(r"第四课堂", str(ws2.cell(row = i, column = 1).value))
    ret3 = re.search(r"国际化模块", str(ws2.cell(row = i, column = 1).value))

    if ret0 != None:
        second = ws2.cell(row = i + 1, column = 5, value = 'P' if (other_courses[0] == 1) else 'N')
    if ret1 != None:
        third = ws2.cell(row = i + 1, column = 5, value = 'P' if (other_courses[1] == 1) else 'N')
    if ret2 != None:
        fourth = ws2.cell(row = i + 1, column = 5, value = 'P' if (other_courses[2] == 1) else 'N')
    if ret3 != None:
        fifth = ws2.cell(row = i + 1, column = 5, value = 'P' if (other_courses[3] == 1) else 'N')

# General Courses
boya = course([], [], [], [], [])
hexin = course([], [], [], [], [])
sixuaner = course([], [], [], [], [])
qita = course([], [], [], [], [])

qita_num = check_and_fill(r"通识选修课程", r"通识核心课程", all, qita)
hexin_num = check_and_fill(r"通识核心课程", r"博雅技艺", all, hexin)
boya_num = check_and_fill(r"博雅技艺", r"中华传统", all, boya)
sixuaner_num = check_and_fill(r"中华传统", r"专业必修", all, sixuaner)

tongshi = 0
for i in range(7, 250):
    ret = re.search(r"通识选修课程", str(ws2.cell(row = i, column = 1).value))
    if ret != None:
        tongshi = i
        break

ws2.unmerge_cells(start_row = tongshi + 1, start_column = 2, end_row = tongshi + 4, end_column = 2)

gcourse = [boya, hexin, sixuaner, qita]
for i in range(4):
    ws2.cell(row = tongshi + i + 1, column = 1, value = ','.join(gcourse[i].id))
    ws2.cell(row = tongshi + i + 1, column = 2, value = ','.join(gcourse[i].name))
    ws2.cell(row = tongshi + i + 1, column = 5, value = ','.join(gcourse[i].grade))

print("另存为(带上后缀名xls/xlsx):")
save = input()
wb2.save(save)
